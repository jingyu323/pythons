from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors, PatternFill

#判断变量类型的函数
def typeof(variate):
    type=None
    if isinstance(variate,int):
        type = "int"
    elif isinstance(variate,str):
        type = "str"
    elif isinstance(variate,float):
        type = "float"
    elif isinstance(variate,list):
        type = "list"
    elif isinstance(variate,tuple):
        type = "tuple"
    elif isinstance(variate,dict):
        type = "dict"
    elif isinstance(variate,set):
        type = "set"
    return type
# 返回变量类型
def getType(variate):
    arr = {"int":"整数","float":"浮点","str":"字符串","list":"列表","tuple":"元组","dict":"字典","set":"集合"}
    vartype = typeof(variate)
    if not (vartype in arr):
        return "未知类型"
    return arr[vartype]

# 加载现有的Excel文件或创建一个新的
wb = load_workbook('data.xlsx')
sheet = wb.active


# 定义一个函数来标记大于5的连续数字
# 定义一个函数来标记连续数字不大于5的单元格
def highlight_cells(sheet):
    running_sum = 0
    for row in sheet.iter_rows():

        for cell in row:
            print(getType(cell.value),cell.value)
            if cell.value is not None and isinstance(cell.value, int) or isinstance(cell.value, float):
                running_sum += cell.value
                if running_sum <= 0.003472:

                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                else:
                    running_sum = 0
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            else:
                running_sum = 0
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')



# 调用函数对工作表进行处理
highlight_cells(sheet)

# 保存工作簿
wb.save('example_modified.xlsx')